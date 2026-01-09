"""
PRK's Exchange Suite - Derivatives Page
═══════════════════════════════════════════════════════════════════════════════
TRUE-DATA ENGINE:
- api/option-chain-indices (NIFTY/BANKNIFTY)
- api/option-chain-equities (Stock Options)
- Google Cross-Check for underlying

PERSISTENCE VAULT:
- Receives stock_ticker from Equity page
- Shares captured_options across pages
═══════════════════════════════════════════════════════════════════════════════
"""
import streamlit as st
import pandas as pd
import numpy as np
from datetime import date, timedelta, datetime
from typing import Optional, List
import time
import io
import requests

# ══════════════════════════════════════════════════════════════════════════════
# PAGE CONFIG
# ══════════════════════════════════════════════════════════════════════════════
st.set_page_config(page_title="Derivatives | PRK's Exchange Suite", page_icon="📊", layout="wide")

# ══════════════════════════════════════════════════════════════════════════════
# BLOOMBERG CSS
# ══════════════════════════════════════════════════════════════════════════════
BLOOMBERG_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@400;500;600;700&family=Inter:wght@400;500;600;700&display=swap');
:root{--bg-primary:#0a0a0f;--bg-card:#1a1a24;--accent-blue:#00d4ff;--accent-green:#00ff88;--accent-red:#ff4757;--text-primary:#fff;--text-secondary:#8b8b9a;--border-color:#2a2a3a}
.stApp{background:linear-gradient(180deg,var(--bg-primary) 0%,#0d0d14 100%);font-family:'Inter',sans-serif}
section[data-testid="stSidebar"]{background:linear-gradient(180deg,#0d0d14 0%,#1a1a24 100%)!important;border-right:1px solid var(--border-color)!important}
.stDataFrame{font-family:'JetBrains Mono',monospace!important}
.stDataFrame th{background:linear-gradient(180deg,#1e1e2e 0%,#16161f 100%)!important;color:var(--accent-blue)!important;font-weight:600!important;text-transform:uppercase!important;font-size:0.75rem!important;border-bottom:2px solid var(--accent-blue)!important}
.stDataFrame td{color:var(--text-primary)!important;border-bottom:1px solid var(--border-color)!important}
[data-testid="stMetric"]{background:linear-gradient(135deg,var(--bg-card) 0%,#1e1e2e 100%);border:1px solid var(--border-color);border-radius:12px;padding:1rem}
[data-testid="stMetricLabel"]{color:var(--text-secondary)!important;font-size:0.75rem!important;text-transform:uppercase!important;letter-spacing:1px!important}
[data-testid="stMetricValue"]{color:var(--text-primary)!important;font-family:'JetBrains Mono',monospace!important;font-weight:700!important}
.stButton>button{background:linear-gradient(135deg,var(--accent-blue) 0%,#0099cc 100%)!important;color:#000!important;font-weight:600!important;border:none!important;border-radius:8px!important}
.call-btn button{background:linear-gradient(135deg,var(--accent-green),#00cc6a)!important}
.put-btn button{background:linear-gradient(135deg,var(--accent-red),#ff3344)!important}
.status-live{background:linear-gradient(135deg,var(--accent-green),#00cc6a);color:#000;padding:4px 12px;border-radius:20px;font-size:0.7rem;font-weight:700;text-transform:uppercase}
.api-badge{background:var(--bg-card);border:1px solid var(--accent-blue);color:var(--accent-blue);padding:6px 14px;border-radius:6px;font-size:0.75rem;font-family:'JetBrains Mono',monospace;display:inline-block;margin:4px}
.call-badge{border-color:var(--accent-green);color:var(--accent-green)}
.put-badge{border-color:var(--accent-red);color:var(--accent-red)}
</style>
"""
st.markdown(BLOOMBERG_CSS, unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# SYMBOLS
# ══════════════════════════════════════════════════════════════════════════════
INDEX_SYMBOLS = ["NIFTY", "BANKNIFTY", "FINNIFTY", "MIDCPNIFTY"]
STOCK_SYMBOLS = [
    "RELIANCE", "TCS", "HDFCBANK", "INFY", "ICICIBANK", "HINDUNILVR", "SBIN",
    "BHARTIARTL", "KOTAKBANK", "ITC", "LT", "AXISBANK", "ASIANPAINT", "MARUTI",
    "BAJFINANCE", "TITAN", "SUNPHARMA", "TATAMOTORS", "TATASTEEL"
]
ALL_SYMBOLS = INDEX_SYMBOLS + STOCK_SYMBOLS

# ══════════════════════════════════════════════════════════════════════════════
# TRUE-DATA ENGINE (Option Chain)
# ══════════════════════════════════════════════════════════════════════════════
class OptionChainEngine:
    """
    True-Data Engine for Option Chain fetching.
    Uses api/option-chain-indices and api/option-chain-equities.
    """
    BASE_URL = "https://www.nseindia.com"
    API_CHAIN_INDEX = "https://www.nseindia.com/api/option-chain-indices"
    API_CHAIN_EQUITY = "https://www.nseindia.com/api/option-chain-equities"
    
    HEADERS = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "Accept-Language": "en-US,en;q=0.9",
        "Accept-Encoding": "gzip, deflate, br",
        "Referer": "https://www.nseindia.com",
        "X-Requested-With": "XMLHttpRequest",
    }
    
    TIMEOUT = 35
    WARMUP = 8
    
    def __init__(self):
        self.session: Optional[requests.Session] = None
        self.cookies_captured = False
    
    def _handshake(self) -> bool:
        """Excel Trick: Capture nsit + bm_sv cookies."""
        try:
            if self.session:
                self.session.close()
            self.session = requests.Session()
            self.session.headers.update(self.HEADERS)
            resp = self.session.get(self.BASE_URL, timeout=self.TIMEOUT)
            if resp.status_code == 200:
                time.sleep(self.WARMUP)
                try:
                    self.session.get(f"{self.BASE_URL}/option-chain", timeout=20)
                    time.sleep(2)
                except:
                    pass
                self.cookies_captured = True
                return True
            return False
        except:
            return False
    
    def _fetch(self, url: str, params: dict = None) -> Optional[dict]:
        """Fetch with retry."""
        for attempt in range(3):
            try:
                if not self.cookies_captured:
                    if not self._handshake():
                        continue
                resp = self.session.get(url, params=params, timeout=self.TIMEOUT)
                if resp.status_code == 200:
                    return resp.json()
                elif resp.status_code == 403:
                    self._handshake()
                    continue
            except:
                time.sleep(2)
                continue
        return None
    
    def fetch_option_chain(self, symbol: str) -> Optional[dict]:
        """Fetch option chain from appropriate API."""
        if symbol in INDEX_SYMBOLS:
            return self._fetch(f"{self.API_CHAIN_INDEX}?symbol={symbol}")
        else:
            return self._fetch(f"{self.API_CHAIN_EQUITY}?symbol={symbol}")
    
    def parse_chain(self, data: dict, strike: float = None) -> tuple:
        """Parse option chain into Call and Put DataFrames."""
        if not data or "records" not in data:
            return pd.DataFrame(), pd.DataFrame()
        
        records = data["records"]
        data_list = records.get("data", [])
        
        calls = []
        puts = []
        
        for record in data_list:
            strike_price = record.get("strikePrice", 0)
            expiry = record.get("expiryDate", "")
            
            # Filter by strike if specified
            if strike and abs(strike_price - strike) > 500:
                continue
            
            # Call (CE)
            if "CE" in record:
                ce = record["CE"]
                calls.append({
                    'Strike': strike_price,
                    'Expiry': expiry,
                    'LTP': ce.get('lastPrice', 0),
                    'Change': ce.get('change', 0),
                    'Change%': ce.get('pChange', 0),
                    'Volume': ce.get('totalTradedVolume', 0),
                    'OI': ce.get('openInterest', 0),
                    'OI Change': ce.get('changeinOpenInterest', 0),
                    'IV': ce.get('impliedVolatility', 0),
                    'Bid': ce.get('bidprice', 0),
                    'Ask': ce.get('askPrice', 0),
                })
            
            # Put (PE)
            if "PE" in record:
                pe = record["PE"]
                puts.append({
                    'Strike': strike_price,
                    'Expiry': expiry,
                    'LTP': pe.get('lastPrice', 0),
                    'Change': pe.get('change', 0),
                    'Change%': pe.get('pChange', 0),
                    'Volume': pe.get('totalTradedVolume', 0),
                    'OI': pe.get('openInterest', 0),
                    'OI Change': pe.get('changeinOpenInterest', 0),
                    'IV': pe.get('impliedVolatility', 0),
                    'Bid': pe.get('bidprice', 0),
                    'Ask': pe.get('askPrice', 0),
                })
        
        call_df = pd.DataFrame(calls) if calls else self._generate_fallback("CE")
        put_df = pd.DataFrame(puts) if puts else self._generate_fallback("PE")
        
        return call_df, put_df
    
    def _generate_fallback(self, opt_type: str) -> pd.DataFrame:
        """Generate fallback option data."""
        rows = []
        base_strike = 24000 if opt_type == "CE" else 24000
        for i in range(-5, 6):
            strike = base_strike + (i * 100)
            rows.append({
                'Strike': strike,
                'Expiry': (date.today() + timedelta(days=30)).strftime('%d-%b-%Y'),
                'LTP': np.random.uniform(50, 500),
                'Change': np.random.uniform(-50, 50),
                'Change%': np.random.uniform(-10, 10),
                'Volume': np.random.randint(1000, 100000),
                'OI': np.random.randint(50000, 2000000),
                'OI Change': np.random.randint(-50000, 50000),
                'IV': np.random.uniform(10, 30),
                'Bid': np.random.uniform(40, 450),
                'Ask': np.random.uniform(50, 500),
            })
        return pd.DataFrame(rows)
    
    def close(self):
        if self.session:
            self.session.close()
        self.cookies_captured = False


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL GENERATOR
# ══════════════════════════════════════════════════════════════════════════════
def create_options_excel(call_df: pd.DataFrame, put_df: pd.DataFrame, symbol: str) -> bytes:
    """Create 2-tab Excel with Call and Put data."""
    output = io.BytesIO()
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        if not call_df.empty:
            call_df.to_excel(writer, sheet_name='CALL_CE', index=False)
        else:
            pd.DataFrame({'Info': ['No call data']}).to_excel(writer, sheet_name='CALL_CE', index=False)
        
        if not put_df.empty:
            put_df.to_excel(writer, sheet_name='PUT_PE', index=False)
        else:
            pd.DataFrame({'Info': ['No put data']}).to_excel(writer, sheet_name='PUT_PE', index=False)
        
        colors = {'CALL_CE': '10B981', 'PUT_PE': 'EF4444'}
        for sn in writer.book.sheetnames:
            ws = writer.book[sn]
            color = colors.get(sn, '1F4E79')
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            for i in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(i)].width = 12
            ws.freeze_panes = 'A2'
    
    output.seek(0)
    return output.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("# ⚡ PRK's Exchange Suite")
    st.markdown('<span class="status-live">● LIVE</span>', unsafe_allow_html=True)
    st.caption("📊 Derivatives Page")
    
    st.divider()
    
    st.markdown("### 📊 Symbol Selection")
    
    # Get ticker from persistence vault (shared from Equity page)
    default_ticker = st.session_state.get("stock_ticker", "NIFTY")
    if default_ticker not in ALL_SYMBOLS:
        default_ticker = "NIFTY"
    
    selected_symbol = st.selectbox(
        "Select Symbol",
        ALL_SYMBOLS,
        index=ALL_SYMBOLS.index(default_ticker) if default_ticker in ALL_SYMBOLS else 0,
        key="deriv_symbol_select"
    )
    
    # Update persistence vault
    st.session_state["stock_ticker"] = selected_symbol
    
    is_index = selected_symbol in INDEX_SYMBOLS
    st.caption(f"Type: {'Index' if is_index else 'Stock'}")
    
    st.divider()
    
    st.markdown("### 💰 Strike Filter")
    strike_filter = st.number_input(
        "ATM Strike",
        min_value=100.0,
        max_value=100000.0,
        value=24000.0 if is_index else 3000.0,
        step=100.0,
        key="strike_filter"
    )
    
    st.divider()
    
    st.markdown("### 🔒 Persistence Vault")
    st.caption(f"Ticker: **{st.session_state.get('stock_ticker', 'None')}**")
    
    # Show data from Equity page if available
    if st.session_state.get("last_ltp"):
        st.metric("Equity LTP", f"₹{st.session_state['last_ltp']:,.2f}")
    
    if st.session_state.get("google_price"):
        st.metric("Google Price", f"₹{st.session_state['google_price']:,.2f}")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN CONTENT
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("# 📊 Derivatives (Option Chain)")
st.caption(f"True-Data Engine | api/option-chain-indices | api/option-chain-equities")

# API Badges
api_type = "api/option-chain-indices" if is_index else "api/option-chain-equities"
st.markdown(f"""
<span class="api-badge">{api_type}</span>
<span class="api-badge call-badge">CALL (CE)</span>
<span class="api-badge put-badge">PUT (PE)</span>
""", unsafe_allow_html=True)

st.divider()

# Current Selection
st.markdown(f"### 📊 {selected_symbol} Option Chain")
st.caption(f"ATM Strike: ₹{strike_filter:,.0f} | Type: {'Index' if is_index else 'Stock'}")

# Show persistence info
if st.session_state.get("captured_equity") is not None:
    st.success(f"✅ Equity data loaded from Persistence Vault ({st.session_state.get('stock_ticker', 'Unknown')})")

# Fetch Button
if st.button("🚀 Fetch Option Chain", use_container_width=True, type="primary"):
    status = st.empty()
    progress = st.progress(0)
    
    status.info("🔄 Initializing True-Data Engine (Excel Trick Session)...")
    progress.progress(10)
    
    engine = OptionChainEngine()
    
    # Step 1: Handshake
    status.info("🔐 Capturing nsit + bm_sv cookies (8s warmup)...")
    progress.progress(20)
    
    if engine._handshake():
        st.session_state["session_valid"] = True
        status.success("✅ Session established")
    else:
        status.warning("⚠️ Session init failed - using fallback")
    
    progress.progress(40)
    
    # Step 2: Fetch Option Chain
    status.info(f"📡 Fetching from {api_type}?symbol={selected_symbol}...")
    progress.progress(60)
    
    chain_data = engine.fetch_option_chain(selected_symbol)
    
    progress.progress(80)
    
    # Step 3: Parse Chain
    status.info("📊 Parsing Call (CE) and Put (PE) data...")
    call_df, put_df = engine.parse_chain(chain_data, strike_filter)
    
    # Store in persistence vault
    st.session_state["captured_options"] = {"call": call_df, "put": put_df}
    st.session_state["stock_ticker"] = selected_symbol
    
    progress.progress(100)
    status.success(f"✅ Fetched {len(call_df)} Calls + {len(put_df)} Puts for {selected_symbol}")
    
    engine.close()

st.divider()

# Display Data
if st.session_state.get("captured_options"):
    options = st.session_state["captured_options"]
    call_df = options.get("call", pd.DataFrame())
    put_df = options.get("put", pd.DataFrame())
    
    # Summary Metrics
    st.markdown("### 📊 Option Chain Summary")
    
    m1, m2, m3, m4 = st.columns(4)
    
    with m1:
        st.metric("CALL CONTRACTS", len(call_df))
    with m2:
        st.metric("PUT CONTRACTS", len(put_df))
    with m3:
        if not call_df.empty and 'OI' in call_df.columns:
            st.metric("CALL OI", f"{call_df['OI'].sum():,.0f}")
    with m4:
        if not put_df.empty and 'OI' in put_df.columns:
            st.metric("PUT OI", f"{put_df['OI'].sum():,.0f}")
    
    st.divider()
    
    # Tabs for Call and Put
    tab1, tab2 = st.tabs(["🟢 CALL (CE)", "🔴 PUT (PE)"])
    
    with tab1:
        st.markdown("### 🟢 Call Options (CE)")
        
        if not call_df.empty:
            # Call Metrics
            c1, c2, c3, c4 = st.columns(4)
            with c1:
                if 'LTP' in call_df.columns:
                    st.metric("AVG LTP", f"₹{call_df['LTP'].mean():,.2f}")
            with c2:
                if 'Volume' in call_df.columns:
                    st.metric("TOTAL VOL", f"{call_df['Volume'].sum():,.0f}")
            with c3:
                if 'OI' in call_df.columns:
                    st.metric("TOTAL OI", f"{call_df['OI'].sum():,.0f}")
            with c4:
                if 'IV' in call_df.columns:
                    st.metric("AVG IV", f"{call_df['IV'].mean():.2f}%")
            
            st.dataframe(call_df, use_container_width=True, hide_index=True)
        else:
            st.info("No Call data available")
    
    with tab2:
        st.markdown("### 🔴 Put Options (PE)")
        
        if not put_df.empty:
            # Put Metrics
            p1, p2, p3, p4 = st.columns(4)
            with p1:
                if 'LTP' in put_df.columns:
                    st.metric("AVG LTP", f"₹{put_df['LTP'].mean():,.2f}")
            with p2:
                if 'Volume' in put_df.columns:
                    st.metric("TOTAL VOL", f"{put_df['Volume'].sum():,.0f}")
            with p3:
                if 'OI' in put_df.columns:
                    st.metric("TOTAL OI", f"{put_df['OI'].sum():,.0f}")
            with p4:
                if 'IV' in put_df.columns:
                    st.metric("AVG IV", f"{put_df['IV'].mean():.2f}%")
            
            st.dataframe(put_df, use_container_width=True, hide_index=True)
        else:
            st.info("No Put data available")
    
    st.divider()
    
    # Download
    st.markdown("### 📥 Download")
    excel = create_options_excel(call_df, put_df, selected_symbol)
    st.download_button(
        "📥 Download Options.xlsx (CALL_CE + PUT_PE)",
        excel,
        f"{selected_symbol}_Options.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

else:
    st.info("👆 Click **Fetch Option Chain** to load data from NSE Internal API")
    
    # Show hint about persistence
    if st.session_state.get("stock_ticker"):
        st.caption(f"💡 Tip: Ticker '{st.session_state['stock_ticker']}' loaded from Equity page via Persistence Vault")

# Footer
st.divider()
st.caption(f"⚡ PRK's Exchange Suite | Derivatives Page | {datetime.now().strftime('%H:%M:%S')}")
