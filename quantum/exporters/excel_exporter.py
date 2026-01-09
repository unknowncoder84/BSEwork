"""
Quantum Market Suite - Excel Exporter

Generates professionally formatted Excel files with merged equity and derivative data.
"""

import io
from datetime import datetime
from typing import Dict, Optional, List
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

from quantum.models import MergedStockData
from quantum.persistence import PersistenceManager


class ExcelExporter:
    """Generates professionally formatted Excel files."""
    
    # Styling constants
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
    
    CELL_ALIGNMENT = Alignment(horizontal="right", vertical="center")
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def __init__(self, persistence_manager: Optional[PersistenceManager] = None):
        """Initialize exporter with optional persistence manager."""
        self.persistence = persistence_manager or PersistenceManager()
    
    def export_to_excel(
        self,
        data: Dict[str, MergedStockData],
        filename: Optional[str] = None
    ) -> bytes:
        """Generate Excel file with company tabs and side-by-side data."""
        if not data:
            raise ValueError("No data to export")
        
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        for symbol, stock_data in data.items():
            ws = wb.create_sheet(title=symbol[:31])  # Excel sheet name limit
            self._populate_worksheet(ws, stock_data)
            self.format_worksheet(ws, stock_data)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Record export in history
        if filename:
            self.persistence.add_export_history(filename)
        
        return output.getvalue()

    def _populate_worksheet(self, ws, stock_data: MergedStockData) -> None:
        """Populate worksheet with data."""
        current_col = 1
        
        # Add equity data
        if stock_data.has_equity:
            ws.cell(row=1, column=current_col, value="EQUITY DATA")
            ws.merge_cells(
                start_row=1, start_column=current_col,
                end_row=1, end_column=current_col + len(stock_data.equity_data.columns) - 1
            )
            
            for r_idx, row in enumerate(dataframe_to_rows(stock_data.equity_data, index=False, header=True), 2):
                for c_idx, value in enumerate(row, current_col):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            
            current_col += len(stock_data.equity_data.columns) + 1
        
        # Add call data
        if stock_data.call_data is not None and not stock_data.call_data.empty:
            ws.cell(row=1, column=current_col, value="CALL OPTIONS")
            ws.merge_cells(
                start_row=1, start_column=current_col,
                end_row=1, end_column=current_col + len(stock_data.call_data.columns) - 1
            )
            
            for r_idx, row in enumerate(dataframe_to_rows(stock_data.call_data, index=False, header=True), 2):
                for c_idx, value in enumerate(row, current_col):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            
            current_col += len(stock_data.call_data.columns) + 1
        
        # Add put data
        if stock_data.put_data is not None and not stock_data.put_data.empty:
            ws.cell(row=1, column=current_col, value="PUT OPTIONS")
            ws.merge_cells(
                start_row=1, start_column=current_col,
                end_row=1, end_column=current_col + len(stock_data.put_data.columns) - 1
            )
            
            for r_idx, row in enumerate(dataframe_to_rows(stock_data.put_data, index=False, header=True), 2):
                for c_idx, value in enumerate(row, current_col):
                    ws.cell(row=r_idx, column=c_idx, value=value)
    
    def format_worksheet(self, ws, data: MergedStockData) -> None:
        """Apply professional formatting to worksheet."""
        # Format section headers (row 1)
        for cell in ws[1]:
            if cell.value:
                cell.font = Font(bold=True, size=12, color="FFFFFF")
                cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
        
        # Format column headers (row 2)
        for cell in ws[2]:
            if cell.value:
                cell.font = self.HEADER_FONT
                cell.fill = self.HEADER_FILL
                cell.alignment = self.HEADER_ALIGNMENT
                cell.border = self.THIN_BORDER
        
        # Format data cells
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
            for cell in row:
                if cell.value is not None:
                    cell.alignment = self.CELL_ALIGNMENT
                    cell.border = self.THIN_BORDER
                    
                    # Format numbers
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00'
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width

    def merge_equity_derivative(
        self,
        equity: Optional[pd.DataFrame],
        calls: Optional[pd.DataFrame],
        puts: Optional[pd.DataFrame]
    ) -> pd.DataFrame:
        """Merge equity and derivative data side-by-side."""
        dfs = []
        
        if equity is not None and not equity.empty:
            equity_copy = equity.copy()
            equity_copy.columns = [f"Equity_{col}" for col in equity_copy.columns]
            dfs.append(equity_copy.reset_index(drop=True))
        
        if calls is not None and not calls.empty:
            calls_copy = calls.copy()
            calls_copy.columns = [f"Call_{col}" for col in calls_copy.columns]
            dfs.append(calls_copy.reset_index(drop=True))
        
        if puts is not None and not puts.empty:
            puts_copy = puts.copy()
            puts_copy.columns = [f"Put_{col}" for col in puts_copy.columns]
            dfs.append(puts_copy.reset_index(drop=True))
        
        if not dfs:
            return pd.DataFrame()
        
        return pd.concat(dfs, axis=1)
    
    def get_worksheet_count(self, excel_bytes: bytes) -> int:
        """Get number of worksheets in Excel file (for testing)."""
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(excel_bytes))
        return len(wb.sheetnames)
    
    def get_worksheet_names(self, excel_bytes: bytes) -> List[str]:
        """Get worksheet names from Excel file (for testing)."""
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(excel_bytes))
        return wb.sheetnames
    
    @staticmethod
    def generate_filename(prefix: str = "quantum_export") -> str:
        """Generate timestamped filename."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{prefix}_{timestamp}.xlsx"
