"""
Property-based tests for Excel generator component.
Feature: bse-derivative-downloader
"""
import pytest
import pandas as pd
import io
from datetime import date, timedelta
from hypothesis import given, settings, strategies as st
from openpyxl import load_workbook

from components.excel_generator import (
    create_excel_file, generate_filename, sanitize_filename
)


# Strategies
company_names = st.text(min_size=1, max_size=30, alphabet=st.characters(whitelist_categories=('L', 'N')))
valid_dates = st.dates(min_value=date(2020, 1, 1), max_value=date(2030, 12, 31))


def create_sample_df(num_rows: int = 5) -> pd.DataFrame:
    """Create sample DataFrame for testing."""
    return pd.DataFrame({
        'Date': [f'0{i+1}-Jan-2024' for i in range(num_rows)],
        'Strike Price': [1000 + i * 10 for i in range(num_rows)],
        'Call LTP': [100.0 + i for i in range(num_rows)],
        'Call OI': [10000 + i * 100 for i in range(num_rows)],
        'Put LTP': [80.0 + i for i in range(num_rows)],
        'Put OI': [8000 + i * 100 for i in range(num_rows)]
    })


class TestExcelFileGeneration:
    """
    Property 10: Excel file contains all merged data
    For any merged DataFrame, the generated Excel file should contain 
    all rows and columns from the DataFrame with proper headers.
    Validates: Requirements 5.1, 5.4
    """
    
    @given(
        num_rows=st.integers(min_value=1, max_value=20),
        company=company_names,
        from_date=valid_dates,
        to_date=valid_dates
    )
    @settings(max_examples=50)
    def test_excel_contains_all_data(self, num_rows, company, from_date, to_date):
        """
        Feature: bse-derivative-downloader, Property 10: Excel file contains all merged data
        Validates: Requirements 5.1, 5.4
        """
        # Ensure valid date range
        if from_date > to_date:
            from_date, to_date = to_date, from_date
        
        # Ensure company name is not empty
        if not company.strip():
            company = "TEST"
        
        # Create sample DataFrame
        df = create_sample_df(num_rows)
        
        # Generate Excel file
        excel_bytes = create_excel_file(df, company, from_date, to_date)
        
        # Verify it's valid Excel
        assert excel_bytes is not None
        assert len(excel_bytes) > 0
        
        # Load and verify content
        workbook = load_workbook(io.BytesIO(excel_bytes))
        worksheet = workbook.active
        
        # Verify data is present (accounting for title rows)
        # Row 1: Title, Row 2: Date range, Row 3: Empty, Row 4: Headers, Row 5+: Data
        data_start_row = 5
        
        # Count data rows
        data_rows = 0
        for row in worksheet.iter_rows(min_row=data_start_row):
            if any(cell.value for cell in row):
                data_rows += 1
        
        # Should have all data rows
        assert data_rows == num_rows, f"Expected {num_rows} data rows, got {data_rows}"
    
    @given(
        company=company_names,
        from_date=valid_dates,
        to_date=valid_dates
    )
    @settings(max_examples=50)
    def test_excel_has_proper_headers(self, company, from_date, to_date):
        """
        Feature: bse-derivative-downloader, Property 10: Excel file contains all merged data
        Verify Excel file has proper column headers.
        Validates: Requirements 5.1, 5.4
        """
        if from_date > to_date:
            from_date, to_date = to_date, from_date
        
        if not company.strip():
            company = "TEST"
        
        df = create_sample_df(5)
        excel_bytes = create_excel_file(df, company, from_date, to_date)
        
        workbook = load_workbook(io.BytesIO(excel_bytes))
        worksheet = workbook.active
        
        # Headers are in row 4
        header_row = 4
        headers = [cell.value for cell in worksheet[header_row] if cell.value]
        
        expected_headers = ['Date', 'Strike Price', 'Call LTP', 'Call OI', 'Put LTP', 'Put OI']
        
        for expected in expected_headers:
            assert expected in headers, f"Missing header: {expected}"


class TestFilenameGeneration:
    """
    Property 11: Filename includes context information
    For any company name and date range, the generated Excel filename 
    should contain the company name, from date, and to date in a readable format.
    Validates: Requirements 5.5
    """
    
    @given(
        company=company_names,
        from_date=valid_dates,
        to_date=valid_dates
    )
    @settings(max_examples=100)
    def test_filename_includes_context(self, company, from_date, to_date):
        """
        Feature: bse-derivative-downloader, Property 11: Filename includes context information
        Validates: Requirements 5.5
        """
        # Ensure valid inputs
        if from_date > to_date:
            from_date, to_date = to_date, from_date
        
        if not company.strip():
            company = "TEST"
        
        filename = generate_filename(company, from_date, to_date)
        
        # Verify filename structure
        assert filename.endswith('.xlsx'), "Filename should end with .xlsx"
        assert 'BSE_Data_' in filename, "Filename should contain 'BSE_Data_'"
        
        # Verify dates are in filename
        from_str = from_date.strftime("%Y%m%d")
        to_str = to_date.strftime("%Y%m%d")
        
        assert from_str in filename, f"From date {from_str} should be in filename"
        assert to_str in filename, f"To date {to_str} should be in filename"
        
        # Verify company name (sanitized) is in filename
        sanitized_company = sanitize_filename(company)
        if sanitized_company:
            assert sanitized_company in filename or company.replace(' ', '_')[:20] in filename
    
    @given(
        from_date=valid_dates,
        to_date=valid_dates
    )
    @settings(max_examples=50)
    def test_filename_handles_special_characters(self, from_date, to_date):
        """
        Test that special characters in company name are handled.
        """
        special_companies = [
            "Company/Name",
            "Stock:Symbol",
            "Test<>Name",
            "Name|With|Pipes",
            "Name?Query"
        ]
        
        for company in special_companies:
            filename = generate_filename(company, from_date, to_date)
            
            # Should not contain invalid filename characters
            invalid_chars = ['<', '>', ':', '"', '/', '\\', '|', '?', '*']
            for char in invalid_chars:
                assert char not in filename, f"Filename should not contain '{char}'"


class TestSanitizeFilename:
    """Tests for filename sanitization."""
    
    def test_removes_invalid_characters(self):
        """Test that invalid characters are removed."""
        test_cases = [
            ("test<>file", "testfile"),
            ("file:name", "filename"),
            ("path/to/file", "pathtofile"),
            ("file|name", "filename"),
            ("file?name", "filename"),
            ("file*name", "filename"),
        ]
        
        for input_text, expected_base in test_cases:
            result = sanitize_filename(input_text)
            # Should not contain invalid characters
            for char in '<>:"/\\|?*':
                assert char not in result
    
    def test_replaces_spaces(self):
        """Test that spaces are replaced with underscores."""
        result = sanitize_filename("test file name")
        assert " " not in result
        assert "_" in result
    
    def test_limits_length(self):
        """Test that filename length is limited."""
        long_name = "a" * 100
        result = sanitize_filename(long_name)
        assert len(result) <= 50
