"""
Excel file generator for BSE derivative data.
"""
import io
import re
from datetime import date
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


def sanitize_filename(text: str) -> str:
    """
    Sanitize text for use in filename.
    """
    # Remove or replace invalid characters
    sanitized = re.sub(r'[<>:"/\\|?*]', '', text)
    # Replace spaces with underscores
    sanitized = sanitized.replace(' ', '_')
    # Limit length
    return sanitized[:50]


def generate_filename(company_name: str, from_date: date, to_date: date) -> str:
    """
    Generate meaningful filename including company name and date range.
    
    Args:
        company_name: Name of the company/stock
        from_date: Start date of data range
        to_date: End date of data range
        
    Returns:
        Filename string like "BSE_Data_RELIANCE_20240101_20240131.xlsx"
    """
    sanitized_company = sanitize_filename(company_name)
    from_str = from_date.strftime("%Y%m%d")
    to_str = to_date.strftime("%Y%m%d")
    
    return f"BSE_Data_{sanitized_company}_{from_str}_{to_str}.xlsx"


def apply_excel_formatting(workbook: Workbook, worksheet) -> None:
    """
    Apply professional formatting to Excel worksheet.
    
    Args:
        workbook: Openpyxl Workbook object
        worksheet: Worksheet to format
    """
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    cell_alignment = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Alternating row colors
    light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Format header row
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Format data rows
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
        for cell in row:
            cell.alignment = cell_alignment
            cell.border = thin_border
            
            # Alternating row colors
            if row_idx % 2 == 0:
                cell.fill = light_fill
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 30)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze header row
    worksheet.freeze_panes = 'A2'


def create_excel_file(df: pd.DataFrame, company_name: str, 
                      from_date: date, to_date: date) -> bytes:
    """
    Create formatted Excel file from DataFrame.
    
    Args:
        df: DataFrame with merged derivative data
        company_name: Name of the company/stock
        from_date: Start date of data range
        to_date: End date of data range
        
    Returns:
        Excel file as bytes
    """
    # Create workbook
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Derivative Data"
    
    # Add title row
    worksheet.merge_cells('A1:F1')
    title_cell = worksheet['A1']
    title_cell.value = f"BSE Derivative Data - {company_name}"
    title_cell.font = Font(bold=True, size=14, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add date range info
    worksheet.merge_cells('A2:F2')
    date_cell = worksheet['A2']
    date_cell.value = f"Date Range: {from_date.strftime('%d-%b-%Y')} to {to_date.strftime('%d-%b-%Y')}"
    date_cell.font = Font(italic=True, size=10)
    date_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add empty row
    worksheet.append([])
    
    # Add DataFrame data starting from row 4
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=4):
        for c_idx, value in enumerate(row, start=1):
            worksheet.cell(row=r_idx, column=c_idx, value=value)
    
    # Apply formatting (adjust for title rows)
    # Format header row (row 4)
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    cell_alignment = Alignment(horizontal="center", vertical="center")
    
    for cell in worksheet[4]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Format data rows
    for row_idx in range(5, worksheet.max_row + 1):
        for cell in worksheet[row_idx]:
            cell.alignment = cell_alignment
            cell.border = thin_border
            if row_idx % 2 == 1:
                cell.fill = light_fill
    
    # Auto-adjust column widths (skip merged cells)
    from openpyxl.utils import get_column_letter
    for col_idx in range(1, worksheet.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        
        for row_idx in range(1, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            try:
                if cell.value and not isinstance(cell, type(None)):
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        adjusted_width = min(max(max_length + 2, 10), 30)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze header row
    worksheet.freeze_panes = 'A5'
    
    # Save to bytes
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    
    return output.getvalue()


def create_simple_excel(df: pd.DataFrame) -> bytes:
    """
    Create simple Excel file without extra formatting.
    
    Args:
        df: DataFrame to export
        
    Returns:
        Excel file as bytes
    """
    output = io.BytesIO()
    df.to_excel(output, index=False, engine='openpyxl')
    output.seek(0)
    return output.getvalue()


def create_multi_stock_excel(stock_data: dict, from_date: date, to_date: date) -> bytes:
    """
    Create Excel file with multiple sheets, one per stock.
    
    Args:
        stock_data: Dictionary mapping stock names to DataFrames
        from_date: Start date of data range
        to_date: End date of data range
        
    Returns:
        Excel file as bytes with multiple sheets
    """
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Create summary sheet first
        summary_data = []
        for stock_name, df in stock_data.items():
            if df is not None and not df.empty:
                call_records = len(df[df.get('Call LTP', 'N/A') != 'N/A']) if 'Call LTP' in df.columns else 0
                put_records = len(df[df.get('Put LTP', 'N/A') != 'N/A']) if 'Put LTP' in df.columns else 0
                summary_data.append({
                    'Stock': stock_name,
                    'Total Records': len(df),
                    'Call Records': call_records,
                    'Put Records': put_records,
                    'Date Range': f"{from_date.strftime('%d-%b-%Y')} to {to_date.strftime('%d-%b-%Y')}"
                })
        
        if summary_data:
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        # Create individual sheets for each stock
        for stock_name, df in stock_data.items():
            if df is not None and not df.empty:
                # Sanitize sheet name (max 31 chars, no special chars)
                sheet_name = sanitize_sheet_name(stock_name)
                df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    output.seek(0)
    
    # Now apply formatting using openpyxl
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    
    workbook = load_workbook(output)
    
    # Format each sheet
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        format_worksheet(worksheet)
    
    # Save formatted workbook
    final_output = io.BytesIO()
    workbook.save(final_output)
    final_output.seek(0)
    
    return final_output.getvalue()


def sanitize_sheet_name(name: str) -> str:
    """
    Sanitize string for use as Excel sheet name.
    Sheet names must be <= 31 chars and cannot contain: \\ / * ? : [ ]
    """
    # Remove invalid characters
    invalid_chars = ['\\', '/', '*', '?', ':', '[', ']']
    sanitized = name
    for char in invalid_chars:
        sanitized = sanitized.replace(char, '')
    
    # Limit to 31 characters
    return sanitized[:31]


def format_worksheet(worksheet) -> None:
    """Apply professional formatting to a worksheet."""
    from openpyxl.utils import get_column_letter
    
    if worksheet.max_row < 1:
        return
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Format header row
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Format data rows
    for row_idx in range(2, worksheet.max_row + 1):
        for cell in worksheet[row_idx]:
            cell.alignment = cell_alignment
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = light_fill
    
    # Auto-adjust column widths
    for col_idx in range(1, worksheet.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        
        for row_idx in range(1, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        adjusted_width = min(max(max_length + 2, 10), 30)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze header row
    worksheet.freeze_panes = 'A2'


def generate_multi_stock_filename(stocks: list, from_date: date, to_date: date) -> str:
    """
    Generate filename for multi-stock export.
    
    Args:
        stocks: List of stock names
        from_date: Start date
        to_date: End date
        
    Returns:
        Filename string
    """
    from_str = from_date.strftime("%Y%m%d")
    to_str = to_date.strftime("%Y%m%d")
    
    if len(stocks) == 1:
        return f"BSE_Data_{sanitize_filename(stocks[0])}_{from_str}_{to_str}.xlsx"
    elif len(stocks) <= 3:
        stock_str = "_".join([sanitize_filename(s)[:10] for s in stocks])
        return f"BSE_Data_{stock_str}_{from_str}_{to_str}.xlsx"
    else:
        return f"BSE_MultiStock_{len(stocks)}stocks_{from_str}_{to_str}.xlsx"
